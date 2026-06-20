from datetime import date
from decimal import Decimal
from io import BytesIO

from projects.permissions import owner_required
from django.db.models import Prefetch, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font

from projects.company import get_company_info
from projects.models import Expense, Income, OperationalExpense, OperationalIncome, Project, Quote, QuoteLineItem
from projects.report_period import parse_report_period

from .pdf import render_pdf


@owner_required
def export_monthly_excel(request):
    period = parse_report_period(request)
    start = period["start"]
    end = period["end"]
    year = period["year"]
    month = period["month"]
    period_label = period["period_label"]
    period_slug = period["period_slug"]

    wb = Workbook()

    ws_income = wb.active
    ws_income.title = "Έσοδα"
    ws_income.append(["Έργο", "Ημερομηνία", "Ποσό", "Τύπος", "Πληρωμή", "Περιγραφή"])
    for cell in ws_income[1]:
        cell.font = Font(bold=True)

    incomes = Income.objects.filter(date__gte=start, date__lt=end).select_related(
        "project", "income_type", "payment_method"
    )
    for income in incomes:
        ws_income.append(
            [
                income.project.name,
                income.date.isoformat(),
                float(income.amount),
                income.income_type.label,
                income.payment_method.label,
                income.description,
            ]
        )

    ws_expense = wb.create_sheet("Εξοδα εργων")
    ws_expense.append(["Εργο", "Ημερομηνία", "Ποσό", "Κατηγορία", "Προμηθευτής", "Περιγραφή"])
    for cell in ws_expense[1]:
        cell.font = Font(bold=True)

    expenses = Expense.objects.filter(date__gte=start, date__lt=end).select_related(
        "project", "category"
    )
    for expense in expenses:
        ws_expense.append(
            [
                expense.project.name,
                expense.date.isoformat(),
                float(expense.amount),
                expense.category.label,
                expense.supplier,
                expense.description,
            ]
        )

    ws_operational = wb.create_sheet("Λειτουργικα εξοδα")
    ws_operational.append(["Ημερομηνία", "Ποσό", "Κατηγορία", "Προμηθευτής", "Περιγραφή"])
    for cell in ws_operational[1]:
        cell.font = Font(bold=True)

    operational = OperationalExpense.objects.filter(date__gte=start, date__lt=end).select_related(
        "category"
    )
    for item in operational:
        ws_operational.append(
            [
                item.date.isoformat(),
                float(item.amount),
                item.category.label,
                item.supplier,
                item.description,
            ]
        )

    ws_operational_income = wb.create_sheet("Λειτουργικα εσοδα")
    ws_operational_income.append(["Ημερομηνία", "Ποσό", "Κατηγορία", "Πηγή", "Περιγραφή"])
    for cell in ws_operational_income[1]:
        cell.font = Font(bold=True)

    operational_incomes = OperationalIncome.objects.filter(
        date__gte=start, date__lt=end
    ).select_related("category")
    for item in operational_incomes:
        ws_operational_income.append(
            [
                item.date.isoformat(),
                float(item.amount),
                item.category.label,
                item.source,
                item.description,
            ]
        )

    total_income = incomes.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
    total_operational_incomes = operational_incomes.aggregate(total=Sum("amount"))[
        "total"
    ] or Decimal("0.00")
    total_project_expenses = expenses.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
    total_operational = operational.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
    total_expenses = total_project_expenses + total_operational
    total_all_income = total_income + total_operational_incomes

    ws_summary = wb.create_sheet("Σύνοψη")
    ws_summary.append(["Περίοδος", period_label])
    ws_summary.append(["Έσοδα έργων", float(total_income)])
    ws_summary.append(["Λειτουργικά έσοδα", float(total_operational_incomes)])
    ws_summary.append(["Σύνολο εσόδων", float(total_all_income)])
    ws_summary.append(["Έξοδα έργων", float(total_project_expenses)])
    ws_summary.append(["Λειτουργικά έξοδα", float(total_operational)])
    ws_summary.append(["Σύνολο εξόδων", float(total_expenses)])
    ws_summary.append(["Καθαρό αποτέλεσμα", float(total_all_income - total_expenses)])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="anafora_{period_slug}.xlsx"'
    return response


@owner_required
def export_project_pdf(request, pk):
    project = get_object_or_404(Project, pk=pk)
    html = render_to_string(
        "reports/project_pdf.html",
        {
            "project": project,
            "incomes": project.incomes.select_related("income_type", "payment_method"),
            "expenses": project.expenses.select_related("category"),
            "work_hours_entries": project.work_hours.select_related("worker", "worker__profile").all(),
            "generated_at": timezone.localtime(),
        },
    )

    try:
        pdf = render_pdf(html, base_url=request.build_absolute_uri("/"))
    except Exception:
        return HttpResponse(
            "Αποτυχία δημιουργίας PDF. Δοκίμασε ξανά.",
            status=503,
        )

    filename = f"ergo_{project.pk}.pdf"
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@owner_required
def export_monthly_pdf(request):
    period = parse_report_period(request)
    start = period["start"]
    end = period["end"]
    period_label = period["period_label"]
    period_slug = period["period_slug"]

    month_incomes = Income.objects.filter(date__gte=start, date__lt=end).select_related(
        "project", "income_type", "payment_method"
    )
    month_project_expenses = Expense.objects.filter(date__gte=start, date__lt=end).select_related(
        "project", "category"
    )
    month_operational_expenses = OperationalExpense.objects.filter(
        date__gte=start, date__lt=end
    ).select_related("category")
    month_operational_incomes = OperationalIncome.objects.filter(
        date__gte=start, date__lt=end
    ).select_related("category")
    total_income = month_incomes.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
    total_operational_incomes = month_operational_incomes.aggregate(total=Sum("amount"))[
        "total"
    ] or Decimal("0.00")
    total_project_expenses = month_project_expenses.aggregate(total=Sum("amount"))[
        "total"
    ] or Decimal("0.00")
    total_operational_expenses = month_operational_expenses.aggregate(
        total=Sum("amount")
    )["total"] or Decimal("0.00")
    total_expenses = total_project_expenses + total_operational_expenses
    total_all_income = total_income + total_operational_incomes

    html = render_to_string(
        "reports/monthly_pdf.html",
        {
            "month_label": period_label,
            "month_incomes": month_incomes,
            "month_project_expenses": month_project_expenses,
            "month_operational_expenses": month_operational_expenses,
            "month_operational_incomes": month_operational_incomes,
            "total_income": total_income,
            "total_operational_incomes": total_operational_incomes,
            "total_all_income": total_all_income,
            "total_project_expenses": total_project_expenses,
            "total_operational_expenses": total_operational_expenses,
            "total_expenses": total_expenses,
            "total_profit": total_all_income - total_expenses,
            "generated_at": timezone.localtime(),
        },
    )

    try:
        pdf = render_pdf(html, base_url=request.build_absolute_uri("/"))
    except Exception:
        return HttpResponse(
            "Αποτυχία δημιουργίας PDF. Δοκίμασε ξανά.",
            status=503,
        )

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="anafora_{period_slug}.pdf"'
    return response


@owner_required
def export_quote_pdf(request, pk):
    from reports.quote_table_layout import get_quote_pdf_table_layout

    quote = get_object_or_404(
        Quote.objects.prefetch_related(
            Prefetch(
                "line_items",
                queryset=QuoteLineItem.objects.select_related("category", "unit"),
            )
        ),
        pk=pk,
    )
    html = render_to_string(
        "reports/quote_pdf.html",
        {
            "quote": quote,
            "line_items": quote.line_items.all(),
            "quote_table": get_quote_pdf_table_layout(),
            "company": get_company_info(),
            "generated_at": timezone.localtime(),
        },
    )

    try:
        pdf = render_pdf(html, base_url=request.build_absolute_uri("/"))
    except Exception:
        return HttpResponse(
            "Αποτυχία δημιουργίας PDF.",
            status=503,
        )

    filename = f"Prosfora_{quote.quote_number}.pdf"
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
